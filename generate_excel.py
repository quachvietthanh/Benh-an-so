import csv
import os

csv_path = r"c:\duanchinh\Benh-an-so\Benh-an-so\Bao_Cao_Kiem_Thu_No_Show_ABC-158.csv"
xlsx_path = r"c:\duanchinh\Benh-an-so\Benh-an-so\Bao_Cao_Kiem_Thu_No_Show_ABC-158.xlsx"

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kiem Thu No-Show ABC-158"

    # Header style
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    # Read CSV
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            ws.append(row)
            if row_idx == 1:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                for cell_idx, cell in enumerate(ws[row_idx], 1):
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if cell.value == "PASSED":
                        cell.font = Font(name="Calibri", size=11, bold=True, color="006100")
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    wb.save(xlsx_path)
    print(f"Successfully generated {xlsx_path}")
except Exception as e:
    print(f"Note: CSV exported at {csv_path}. Openpyxl status: {e}")
